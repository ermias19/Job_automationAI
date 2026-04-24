from __future__ import annotations

import csv
import logging
from pathlib import Path
import re
import socket
from typing import Any

from job_automation.config import Settings
from job_automation.defaults import DEFAULT_PHD_REPORT_HEADERS, DEFAULT_SHEET_HEADERS
from job_automation.models import JobListing, MatchResult
from job_automation.reports import build_job_automation_rows, build_phd_role_rows

logger = logging.getLogger(__name__)


class SheetExporter:
    """Coordinator that exports two independent report streams:
    1) Jobs sheet
    2) PhD roles sheet
    """

    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.job_exporter = JobAutomationWorksheetExporter(settings)
        self.remote_job_exporter = RemoteJobsWorksheetExporter(settings)
        self.phd_exporter = PhdRoleWorksheetExporter(settings)

    def export(
        self,
        run_id: str,
        searched_at: str,
        matches: list[MatchResult],
        output_dir: Path,
    ) -> dict:
        onsite_matches: list[MatchResult] = []
        remote_matches: list[MatchResult] = []
        for match in matches:
            if self._is_remote_job(match.job):
                remote_matches.append(match)
            else:
                onsite_matches.append(match)

        job_rows = build_job_automation_rows(
            run_id=run_id,
            searched_at=searched_at,
            matches=onsite_matches,
        )
        csv_path = output_dir / "matched_jobs.csv"
        self._write_local_csv(csv_path, job_rows, DEFAULT_SHEET_HEADERS)
        xlsx_path = output_dir / "matched_jobs.xlsx"
        xlsx_written = self._write_local_xlsx(
            xlsx_path,
            job_rows,
            DEFAULT_SHEET_HEADERS,
            worksheet_title="job-automation",
        )

        remote_job_rows = build_job_automation_rows(
            run_id=run_id,
            searched_at=searched_at,
            matches=remote_matches,
        )
        remote_csv_path = output_dir / "matched_remote_jobs.csv"
        self._write_local_csv(remote_csv_path, remote_job_rows, DEFAULT_SHEET_HEADERS)
        remote_xlsx_path = output_dir / "matched_remote_jobs.xlsx"
        remote_xlsx_written = self._write_local_xlsx(
            remote_xlsx_path,
            remote_job_rows,
            DEFAULT_SHEET_HEADERS,
            worksheet_title="remote-jobs",
        )

        jobs_remote_status = "skipped"
        if job_rows:
            jobs_remote_status = self._export_jobs_only_remote(
                exporter=self.job_exporter,
                rows=job_rows,
                export_label="Jobs",
            )

        remote_jobs_remote_status = "skipped"
        if remote_job_rows:
            remote_jobs_remote_status = self._export_jobs_only_remote(
                exporter=self.remote_job_exporter,
                rows=remote_job_rows,
                export_label="Remote Jobs",
            )

        remote_status = jobs_remote_status
        if remote_jobs_remote_status != "skipped":
            if jobs_remote_status == "skipped":
                remote_status = f"remote_jobs:{remote_jobs_remote_status}"
            elif jobs_remote_status == remote_jobs_remote_status:
                remote_status = jobs_remote_status
            else:
                remote_status = f"jobs:{jobs_remote_status};remote_jobs:{remote_jobs_remote_status}"

        return {
            "csv_path": str(csv_path),
            "xlsx_path": str(xlsx_path) if xlsx_written else "",
            "remote_jobs_csv_path": str(remote_csv_path),
            "remote_jobs_xlsx_path": str(remote_xlsx_path) if remote_xlsx_written else "",
            "phd_report_csv_path": "",
            "phd_report_xlsx_path": "",
            "jobs_remote_status": jobs_remote_status,
            "remote_jobs_remote_status": remote_jobs_remote_status,
            "remote_status": remote_status,
        }

    def export_phd_only(
        self,
        run_id: str,
        searched_at: str,
        matches: list[MatchResult],
        output_dir: Path,
    ) -> dict:
        phd_rows = build_phd_role_rows(
            run_id=run_id,
            searched_at=searched_at,
            matches=matches,
        )
        phd_report_csv_path = output_dir / "phd_research_report.csv"
        self._write_local_csv(phd_report_csv_path, phd_rows, DEFAULT_PHD_REPORT_HEADERS)
        phd_report_xlsx_path = output_dir / "phd_research_report.xlsx"
        xlsx_written = self._write_local_xlsx(
            phd_report_xlsx_path,
            phd_rows,
            DEFAULT_PHD_REPORT_HEADERS,
            worksheet_title="phd-research-report",
        )

        # Safeguard: PhD report writes are full refreshes so stale rows never accumulate.
        remote_status = self._export_phd_only_remote(phd_rows)

        return {
            "csv_path": "",
            "xlsx_path": "",
            "phd_report_csv_path": str(phd_report_csv_path),
            "phd_report_xlsx_path": str(phd_report_xlsx_path) if xlsx_written else "",
            "remote_status": remote_status,
        }

    def _export_jobs_only_remote(
        self,
        exporter: BaseWorksheetExporter,
        rows: list[dict],
        export_label: str,
    ) -> str:
        can_use_service_account = (
            self.settings.google_sheets_spreadsheet_id
            and self.settings.google_service_account_json
            and self.settings.google_service_account_json.exists()
        )
        can_use_apps_script = bool(self.settings.google_apps_script_webapp_url)

        if (can_use_service_account or can_use_apps_script) and self._google_dns_unavailable():
            logger.warning(
                "Google DNS/network unavailable; skipping remote %s export for this run.",
                export_label,
            )
            return "error:NetworkUnavailable"

        if can_use_service_account:
            try:
                exporter.append_via_service_account(rows)
                return "service_account"
            except Exception as exc:
                if self._is_dns_resolution_error(exc):
                    logger.warning(
                        "Service-account %s export skipped due to DNS/network issue: %s",
                        export_label,
                        exc,
                    )
                    return "error:NetworkUnavailable"
                logger.exception(
                    "Service-account %s export failed; trying Apps Script fallback",
                    export_label,
                )
                if can_use_apps_script:
                    try:
                        exporter.append_via_apps_script(rows)
                        return "apps_script_fallback"
                    except Exception as exc:
                        if self._is_dns_resolution_error(exc):
                            logger.warning(
                                "Apps Script %s export skipped due to DNS/network issue: %s",
                                export_label,
                                exc,
                            )
                            return "error:NetworkUnavailable"
                        logger.exception(
                            "Apps Script fallback also failed for %s; local CSV was still written",
                            export_label,
                        )
                        return f"error:{exc.__class__.__name__}"
                return "error:ServiceAccountExportFailed"

        if can_use_apps_script:
            try:
                exporter.append_via_apps_script(rows)
                return "apps_script"
            except Exception as exc:
                if self._is_dns_resolution_error(exc):
                    logger.warning(
                        "Apps Script %s export skipped due to DNS/network issue: %s",
                        export_label,
                        exc,
                    )
                    return "error:NetworkUnavailable"
                logger.exception(
                    "Apps Script export failed for %s; local CSV was still written",
                    export_label,
                )
                return f"error:{exc.__class__.__name__}"

        return "skipped"

    @staticmethod
    def _is_remote_job(job: JobListing) -> bool:
        if bool((job.raw or {}).get("search_target_remote")):
            return True

        haystack = " ".join(
            [
                job.job_location or "",
                job.job_title or "",
                job.job_summary or "",
                job.job_description_formatted or "",
            ]
        ).lower()

        remote_markers = (
            "remote-",
            "remote,",
            "work from home",
            " wfh ",
            "fully remote",
            "100% remote",
            "anywhere",
        )
        padded = f" {haystack} "
        if any(marker in padded for marker in remote_markers):
            return True
        return re.search(r"\bremote\b", padded) is not None

    def _export_phd_only_remote(self, phd_rows: list[dict]) -> str:
        can_use_service_account = (
            self.settings.google_sheets_spreadsheet_id
            and self.settings.google_service_account_json
            and self.settings.google_service_account_json.exists()
        )
        can_use_apps_script = bool(self.settings.google_apps_script_webapp_url)

        if (can_use_service_account or can_use_apps_script) and self._google_dns_unavailable():
            logger.warning("Google DNS/network unavailable; skipping remote PhD export for this run.")
            return "error:NetworkUnavailable"

        if can_use_service_account:
            try:
                self.phd_exporter.replace_via_service_account(phd_rows)
                return "service_account"
            except Exception as exc:
                if self._is_dns_resolution_error(exc):
                    logger.warning(
                        "Service-account PhD export skipped due to DNS/network issue: %s",
                        exc,
                    )
                    return "error:NetworkUnavailable"
                logger.exception("Service-account export failed for PhD sheet; trying Apps Script fallback")
                if can_use_apps_script:
                    try:
                        self.phd_exporter.replace_via_apps_script(phd_rows)
                        return "apps_script_fallback"
                    except Exception as exc:
                        if self._is_dns_resolution_error(exc):
                            logger.warning(
                                "Apps Script PhD export skipped due to DNS/network issue: %s",
                                exc,
                            )
                            return "error:NetworkUnavailable"
                        logger.exception("Apps Script fallback also failed for PhD sheet")
                        return f"error:{exc.__class__.__name__}"
                return "error:ServiceAccountExportFailed"

        if can_use_apps_script:
            try:
                self.phd_exporter.replace_via_apps_script(phd_rows)
                return "apps_script"
            except Exception as exc:
                if self._is_dns_resolution_error(exc):
                    logger.warning(
                        "Apps Script PhD export skipped due to DNS/network issue: %s",
                        exc,
                    )
                    return "error:NetworkUnavailable"
                logger.exception("Apps Script export failed for PhD sheet")
                return f"error:{exc.__class__.__name__}"

        return "skipped"

    @staticmethod
    def _google_dns_unavailable() -> bool:
        hosts = ("oauth2.googleapis.com", "script.google.com")
        for host in hosts:
            try:
                socket.getaddrinfo(host, 443)
                return False
            except OSError:
                continue
        return True

    @staticmethod
    def _is_dns_resolution_error(exc: Exception) -> bool:
        markers = (
            "name resolution",
            "name or service not known",
            "failed to resolve",
            "noderesolutionerror",
            "nodename nor servname provided",
            "temporary failure in name resolution",
        )
        current: BaseException | None = exc
        while current is not None:
            text = str(current).lower()
            if any(marker in text for marker in markers):
                return True
            current = current.__cause__
        return False

    @staticmethod
    def _write_local_csv(path: Path, rows: list[dict], headers: list[str]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)

    @staticmethod
    def _write_local_xlsx(
        path: Path,
        rows: list[dict],
        headers: list[str],
        worksheet_title: str,
    ) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.warning("openpyxl is not installed; skipping XLSX export for %s", path.name)
            return False

        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = worksheet_title[:31] if worksheet_title else "report"

        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.fill = PatternFill(fill_type="solid", fgColor="339966")
            cell.font = Font(color="FFFFFF", bold=True)

        for row_index, row in enumerate(rows, start=2):
            for col_index, header in enumerate(headers, start=1):
                value: Any = row.get(header, "")
                sheet.cell(row=row_index, column=col_index, value=value)

        workbook.save(path)
        return True


class BaseWorksheetExporter:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    @property
    def worksheet_name(self) -> str:
        raise NotImplementedError

    @property
    def headers(self) -> list[str]:
        raise NotImplementedError

    def append_via_apps_script(self, rows: list[dict]) -> None:
        import requests

        response = requests.post(
            self.settings.google_apps_script_webapp_url,
            json={
                "spreadsheetId": self.settings.google_sheets_spreadsheet_id,
                "worksheet": self.worksheet_name,
                "headers": self.headers,
                "rows": rows,
            },
            timeout=30,
        )
        if response.status_code >= 400:
            raise RuntimeError(
                f"Apps Script export failed with HTTP {response.status_code}: {response.text[:300]}"
            )
        response.raise_for_status()

    def replace_via_apps_script(self, rows: list[dict]) -> None:
        import requests

        response = requests.post(
            self.settings.google_apps_script_webapp_url,
            json={
                "spreadsheetId": self.settings.google_sheets_spreadsheet_id,
                "worksheet": self.worksheet_name,
                "headers": self.headers,
                "rows": rows,
                "clearExisting": True,
            },
            timeout=30,
        )
        if response.status_code >= 400:
            raise RuntimeError(
                f"Apps Script export failed with HTTP {response.status_code}: {response.text[:300]}"
            )
        response.raise_for_status()

    def append_via_service_account(self, rows: list[dict]) -> None:
        import gspread

        client = gspread.service_account(filename=str(self.settings.google_service_account_json))
        spreadsheet = client.open_by_key(self.settings.google_sheets_spreadsheet_id)
        worksheet = self._get_or_create_worksheet(spreadsheet)
        resolved_headers, _ = self._ensure_headers(worksheet)
        self._style_header_row(worksheet, len(resolved_headers))
        self._after_header_ready(worksheet=worksheet, headers=resolved_headers)
        self._append_rows_from_column_a(
            worksheet=worksheet,
            values=[[row.get(header, "") for header in resolved_headers] for row in rows],
        )

    def replace_via_service_account(self, rows: list[dict]) -> None:
        import gspread

        client = gspread.service_account(filename=str(self.settings.google_service_account_json))
        spreadsheet = client.open_by_key(self.settings.google_sheets_spreadsheet_id)
        worksheet = self._get_or_create_worksheet(spreadsheet)
        self._clear_worksheet_data(worksheet)

        resolved_headers, _ = self._ensure_headers(worksheet)
        self._style_header_row(worksheet, len(resolved_headers))
        self._after_header_ready(worksheet=worksheet, headers=resolved_headers)

        if not rows:
            return

        self._append_rows_from_column_a(
            worksheet=worksheet,
            values=[[row.get(header, "") for header in resolved_headers] for row in rows],
        )

    def _after_header_ready(self, worksheet, headers: list[str]) -> None:
        """Hook for subclass-specific worksheet setup after headers are ready."""
        return

    def _get_or_create_worksheet(self, spreadsheet):
        try:
            return spreadsheet.worksheet(self.worksheet_name)
        except Exception:
            return spreadsheet.add_worksheet(
                title=self.worksheet_name,
                rows=1000,
                cols=len(self.headers),
            )

    def _ensure_headers(self, worksheet) -> tuple[list[str], str | None]:
        header_row = self._read_header_row(worksheet)
        existing_headers = [value for value in header_row if value]
        header_start_index = next((idx for idx, value in enumerate(header_row) if value), None)
        if not existing_headers:
            try:
                self._write_headers_at_a1(worksheet, self.headers)
                return list(self.headers), None
            except Exception:
                logger.warning("Header row could not be written; appending data starting from A2")
                return list(self.headers), "A2"

        extra_headers = [header for header in existing_headers if header not in self.headers]
        target_headers = list(self.headers) + extra_headers
        needs_reanchor = header_start_index not in (None, 0)

        if existing_headers == target_headers and not needs_reanchor:
            self._repair_legacy_rows_if_needed(worksheet=worksheet, target_headers=target_headers)
            return existing_headers, None

        try:
            self._migrate_existing_rows_to_target_headers(
                worksheet=worksheet,
                existing_headers=existing_headers,
                target_headers=target_headers,
                source_start_col_index=(header_start_index or 0) + 1,
            )
            self._clear_header_row(
                worksheet=worksheet,
                column_count=max((header_start_index or 0) + len(existing_headers), len(target_headers)),
            )
            self._write_headers_at_a1(worksheet, target_headers)
            return target_headers, None
        except Exception:
            logger.warning("Could not update header row; using existing headers as-is")
            return existing_headers, None

    def _repair_legacy_rows_if_needed(self, worksheet, target_headers: list[str]) -> None:
        """Default: no-op. Subclass can override."""
        return

    def _migrate_existing_rows_to_target_headers(
        self,
        worksheet,
        existing_headers: list[str],
        target_headers: list[str],
        source_start_col_index: int = 1,
    ) -> None:
        used_last_row = worksheet.get_all_values()
        if not used_last_row:
            return

        used_last_row_index = len(used_last_row)
        if used_last_row_index <= 1:
            return

        source_start_col = self._column_index_to_letter(source_start_col_index)
        source_end_col_index = source_start_col_index + len(existing_headers) - 1
        source_end_col = self._column_index_to_letter(source_end_col_index)
        current_rows = worksheet.get(
            f"{source_start_col}2:{source_end_col}{used_last_row_index}",
            value_render_option="FORMULA",
        )
        if not current_rows:
            return

        normalized_rows: list[list[str]] = []
        for row in current_rows:
            row_padded = row + [""] * (len(existing_headers) - len(row))
            row_dict = {header: row_padded[idx] for idx, header in enumerate(existing_headers)}
            normalized_rows.append([row_dict.get(header, "") for header in target_headers])

        new_end_col = self._column_index_to_letter(len(target_headers))
        clear_end_col = self._column_index_to_letter(max(source_end_col_index, len(target_headers)))
        self._ensure_grid_size(
            worksheet=worksheet,
            min_rows=used_last_row_index,
            min_cols=max(source_end_col_index, len(target_headers)),
        )
        worksheet.batch_clear([f"A2:{clear_end_col}{used_last_row_index}"])
        worksheet.update(
            range_name=f"A2:{new_end_col}{used_last_row_index}",
            values=normalized_rows,
            value_input_option="USER_ENTERED",
        )

    def _read_header_row(self, worksheet) -> list[str]:
        probe_columns = max(len(self.headers), getattr(worksheet, "col_count", len(self.headers)))
        end_col = self._column_index_to_letter(probe_columns)
        values = worksheet.get(f"A1:{end_col}1")
        if not values:
            return []
        row = list(values[0])
        if len(row) < probe_columns:
            row.extend([""] * (probe_columns - len(row)))
        return row

    def _write_headers_at_a1(self, worksheet, headers: list[str]) -> None:
        self._ensure_grid_size(worksheet=worksheet, min_rows=1, min_cols=len(headers))
        end_col = self._column_index_to_letter(len(headers))
        worksheet.update(
            range_name=f"A1:{end_col}1",
            values=[headers],
            value_input_option="USER_ENTERED",
        )

    def _clear_header_row(self, worksheet, column_count: int) -> None:
        if column_count <= 0:
            return
        end_col = self._column_index_to_letter(column_count)
        worksheet.batch_clear([f"A1:{end_col}1"])

    def _append_rows_from_column_a(self, worksheet, values: list[list[str]]) -> int:
        if not values:
            return max(len(worksheet.get_all_values()) + 1, 2)

        base_row_count = len(worksheet.get_all_values())
        start_row = max(base_row_count + 1, 2)
        max_cols = max(len(row) for row in values)
        normalized_rows = [row + [""] * (max_cols - len(row)) for row in values]
        end_row = start_row + len(normalized_rows) - 1
        end_col = self._column_index_to_letter(max_cols)

        self._ensure_grid_size(worksheet=worksheet, min_rows=end_row, min_cols=max_cols)
        worksheet.update(
            range_name=f"A{start_row}:{end_col}{end_row}",
            values=normalized_rows,
            value_input_option="USER_ENTERED",
        )
        return start_row

    @staticmethod
    def _ensure_grid_size(worksheet, min_rows: int = 0, min_cols: int = 0) -> None:
        current_rows = getattr(worksheet, "row_count", 0) or 0
        current_cols = getattr(worksheet, "col_count", 0) or 0

        if min_rows > current_rows:
            worksheet.add_rows(min_rows - current_rows)
        if min_cols > current_cols:
            worksheet.add_cols(min_cols - current_cols)

    def _clear_worksheet_data(self, worksheet) -> None:
        try:
            worksheet.clear()
            return
        except Exception:
            logger.warning(
                "Could not fully clear worksheet %s; trying row-content clear fallback",
                self.worksheet_name,
            )
        try:
            worksheet.batch_clear(["A2:ZZZ"])
        except Exception:
            logger.warning(
                "Could not clear worksheet row contents for %s; continuing",
                self.worksheet_name,
            )

    def _style_header_row(self, worksheet, header_count: int) -> None:
        if header_count <= 0:
            return

        end_col = self._column_index_to_letter(header_count)
        header_range = f"A1:{end_col}1"
        try:
            worksheet.format(
                header_range,
                {
                    "backgroundColor": {"red": 0.20, "green": 0.60, "blue": 0.20},
                    "textFormat": {
                        "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                        "bold": True,
                    },
                },
            )
        except Exception:
            logger.warning("Could not style header row; continuing without formatting")

    @staticmethod
    def _column_index_to_letter(index: int) -> str:
        letters: list[str] = []
        while index > 0:
            index, rem = divmod(index - 1, 26)
            letters.append(chr(65 + rem))
        return "".join(reversed(letters))


class JobAutomationWorksheetExporter(BaseWorksheetExporter):
    @property
    def worksheet_name(self) -> str:
        return self.settings.google_sheets_worksheet

    @property
    def headers(self) -> list[str]:
        return DEFAULT_SHEET_HEADERS

    def append_via_service_account(self, rows: list[dict]) -> None:
        import gspread

        if not rows:
            return

        client = gspread.service_account(filename=str(self.settings.google_service_account_json))
        spreadsheet = client.open_by_key(self.settings.google_sheets_spreadsheet_id)
        worksheet = self._get_or_create_worksheet(spreadsheet)
        resolved_headers, _ = self._ensure_headers(worksheet)
        self._style_header_row(worksheet, len(resolved_headers))
        self._after_header_ready(worksheet=worksheet, headers=resolved_headers)

        phase_label = self._next_phase_label(worksheet=worksheet, headers=resolved_headers)
        phase_rows = [self._attach_phase(row=row, phase=phase_label) for row in rows]
        filtered_rows = self._filter_new_rows(
            worksheet=worksheet,
            headers=resolved_headers,
            rows=phase_rows,
        )
        source_rows = filtered_rows or phase_rows

        spacer_row = [""] * len(resolved_headers)
        phase_marker = {header: "" for header in resolved_headers}
        if "Job Title" in resolved_headers:
            if filtered_rows:
                phase_marker["Job Title"] = f"PHASE {phase_label}"
            else:
                phase_marker["Job Title"] = f"PHASE {phase_label} (NO NEW JOBS)"
        if "Phase" in resolved_headers:
            phase_marker["Phase"] = phase_label
        if "Run ID" in resolved_headers:
            phase_marker["Run ID"] = str(source_rows[0].get("Run ID", "")).strip()
        if "Searched At" in resolved_headers:
            phase_marker["Searched At"] = str(source_rows[0].get("Searched At", "")).strip()

        payload = [spacer_row]
        payload.append([phase_marker.get(header, "") for header in resolved_headers])
        if filtered_rows:
            payload.extend([[row.get(header, "") for header in resolved_headers] for row in filtered_rows])

        start_row = self._append_rows_from_column_a(worksheet=worksheet, values=payload)
        phase_marker_row_index = start_row + 1  # spacer row, then marker row
        self._style_phase_marker_cells(
            worksheet=worksheet,
            headers=resolved_headers,
            row_index=phase_marker_row_index,
        )

        logger.info(
            "Jobs sheet appended %s rows in phase %s (input=%s, deduped=%s)",
            len(filtered_rows),
            phase_label,
            len(rows),
            len(rows) - len(filtered_rows),
        )

    @staticmethod
    def _attach_phase(row: dict, phase: str) -> dict:
        enriched = dict(row)
        enriched["Phase"] = phase
        return enriched

    def _next_phase_label(self, worksheet, headers: list[str]) -> str:
        if "Phase" not in headers:
            return "A1"

        phase_col_index = headers.index("Phase") + 1
        values = worksheet.col_values(phase_col_index)
        phase_pattern = re.compile(r"^([ABC])(\d+)$", re.IGNORECASE)

        last_letter = "C"
        last_group = 0
        found = False
        for value in values[1:]:
            match = phase_pattern.match((value or "").strip())
            if not match:
                continue
            last_letter = match.group(1).upper()
            last_group = int(match.group(2))
            found = True

        if not found:
            return "A1"
        if last_letter == "A":
            return f"B{last_group}"
        if last_letter == "B":
            return f"C{last_group}"
        return f"A{last_group + 1}"

    def _filter_new_rows(self, worksheet, headers: list[str], rows: list[dict]) -> list[dict]:
        existing_keys = self._existing_job_keys(worksheet=worksheet, headers=headers)
        if not existing_keys:
            return rows

        filtered: list[dict] = []
        for row in rows:
            key = self._job_key_from_row_dict(row)
            if not key or key in existing_keys:
                continue
            filtered.append(row)
            existing_keys.add(key)
        return filtered

    def _existing_job_keys(self, worksheet, headers: list[str]) -> set[str]:
        values = worksheet.get_all_values()
        if len(values) <= 1:
            return set()

        index_map = {header: idx for idx, header in enumerate(headers)}
        title_idx = index_map.get("Job Title")
        if title_idx is None:
            return set()

        existing: set[str] = set()
        for raw_row in values[1:]:
            row = raw_row + [""] * (len(headers) - len(raw_row))
            title = row[title_idx].strip()
            if title.upper().startswith("PHASE "):
                continue
            key = self._job_key_from_cells(
                apply_link=row[index_map.get("Apply Link", -1)] if "Apply Link" in index_map else "",
                job_title=title,
                company=row[index_map.get("Company", -1)] if "Company" in index_map else "",
                location=row[index_map.get("Location", -1)] if "Location" in index_map else "",
            )
            if key:
                existing.add(key)
        return existing

    def _job_key_from_row_dict(self, row: dict) -> str:
        return self._job_key_from_cells(
            apply_link=str(row.get("Apply Link", "")),
            job_title=str(row.get("Job Title", "")),
            company=str(row.get("Company", "")),
            location=str(row.get("Location", "")),
        )

    @staticmethod
    def _job_key_from_cells(
        apply_link: str,
        job_title: str,
        company: str,
        location: str,
    ) -> str:
        link = apply_link.strip().lower()
        if link:
            return f"link::{link}"
        title = job_title.strip().lower()
        org = company.strip().lower()
        place = location.strip().lower()
        if not (title or org or place):
            return ""
        return f"meta::{title}::{org}::{place}"

    def _style_phase_marker_cells(
        self,
        worksheet,
        headers: list[str],
        row_index: int,
    ) -> None:
        targets: list[str] = []
        if "Job Title" in headers:
            idx = headers.index("Job Title") + 1
            targets.append(f"{self._column_index_to_letter(idx)}{row_index}")
        if "Phase" in headers:
            idx = headers.index("Phase") + 1
            targets.append(f"{self._column_index_to_letter(idx)}{row_index}")

        if not targets:
            return

        fmt = {
            "backgroundColor": {"red": 0.19, "green": 0.45, "blue": 0.85},
            "textFormat": {
                "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                "bold": True,
            },
        }
        try:
            for cell_ref in targets:
                worksheet.format(cell_ref, fmt)
        except Exception:
            logger.warning("Could not style phase marker cells; continuing without phase coloring")

    def _after_header_ready(self, worksheet, headers: list[str]) -> None:
        from gspread.utils import ValidationConditionType

        if "Application Status" not in headers:
            return

        status_col_index = headers.index("Application Status") + 1
        status_col_letter = self._column_index_to_letter(status_col_index)
        status_range = f"{status_col_letter}2:{status_col_letter}"
        try:
            worksheet.add_validation(
                status_range,
                ValidationConditionType.one_of_list,
                ["Applied", "Interview", "Rejected", "Accepted"],
                strict=True,
                showCustomUi=True,
            )
        except Exception:
            logger.warning("Could not apply status dropdown validation; continuing without validation")

        self._apply_status_conditional_formatting(worksheet=worksheet, status_col_index=status_col_index)

    def _apply_status_conditional_formatting(self, worksheet, status_col_index: int) -> None:
        spreadsheet = getattr(worksheet, "spreadsheet", None)
        if spreadsheet is None:
            logger.warning("Could not access spreadsheet object for conditional formatting")
            return

        sheet_id = worksheet.id
        start_col = status_col_index - 1
        end_col = status_col_index

        delete_requests: list[dict] = []
        try:
            metadata = spreadsheet.fetch_sheet_metadata()
            sheets = metadata.get("sheets", [])
            rules: list[dict] = []
            for item in sheets:
                props = item.get("properties", {})
                if props.get("sheetId") == sheet_id:
                    rules = item.get("conditionalFormats", []) or []
                    break

            delete_indexes: list[int] = []
            for idx, rule in enumerate(rules):
                ranges = rule.get("ranges", []) or []
                for range_item in ranges:
                    if (
                        int(range_item.get("startColumnIndex", -1)) == start_col
                        and int(range_item.get("endColumnIndex", -1)) == end_col
                    ):
                        delete_indexes.append(idx)
                        break

            for idx in sorted(delete_indexes, reverse=True):
                delete_requests.append(
                    {
                        "deleteConditionalFormatRule": {
                            "sheetId": sheet_id,
                            "index": idx,
                        }
                    }
                )
        except Exception:
            logger.warning("Could not inspect existing conditional formatting rules; continuing")

        status_range = {
            "sheetId": sheet_id,
            "startRowIndex": 1,  # row 2 onward
            "startColumnIndex": start_col,
            "endColumnIndex": end_col,
        }

        def _text_eq_rule(value: str, fmt: dict) -> dict:
            return {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [status_range],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_EQ",
                                "values": [{"userEnteredValue": value}],
                            },
                            "format": fmt,
                        },
                    },
                    "index": 0,
                }
            }

        add_requests = [
            _text_eq_rule(
                "Applied",
                {
                    "backgroundColor": {"red": 0.20, "green": 0.60, "blue": 0.20},
                    "textFormat": {
                        "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                    },
                },
            ),
            _text_eq_rule(
                "Interview",
                {
                    "backgroundColor": {"red": 1.0, "green": 0.92, "blue": 0.23},
                    "textFormat": {
                        "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                    },
                },
            ),
            _text_eq_rule(
                "Rejected",
                {
                    "backgroundColor": {"red": 0.86, "green": 0.20, "blue": 0.20},
                    "textFormat": {
                        "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                    },
                },
            ),
            _text_eq_rule(
                "Accepted",
                {
                    "textFormat": {
                        "bold": True,
                        "foregroundColor": {"red": 0.10, "green": 0.55, "blue": 0.10},
                    },
                },
            ),
        ]

        requests = delete_requests + add_requests
        try:
            spreadsheet.batch_update({"requests": requests})
        except Exception:
            logger.warning("Could not apply conditional formatting for Application Status column")

    def _repair_legacy_rows_if_needed(self, worksheet, target_headers: list[str]) -> None:
        used_values = worksheet.get_all_values()
        if len(used_values) <= 1:
            return

        yes_no_values = {"Yes", "No"}
        target_len = len(target_headers)
        job_summary_index = target_headers.index("Job Summary")
        app_status_index = target_headers.index("Application Status")
        ai_fit_index = target_headers.index("AI Fit")
        resume_doc_index = target_headers.index("Resume Doc")
        fit_score_index = target_headers.index("Fit Score")

        repaired_rows: list[list[str]] = []
        changed = False

        for row in used_values[1:]:
            row_padded = (row + [""] * target_len)[:target_len]
            app_status_value = row_padded[app_status_index].strip()
            tail_status_value = row_padded[-1].strip()
            ai_fit_value = row_padded[ai_fit_index].strip()
            job_summary_value = row_padded[job_summary_index].strip()
            resume_doc_value = row_padded[resume_doc_index].strip()
            fit_score_value = row_padded[fit_score_index].strip()

            # Case 1: Resume title in Application Status and shift from Resume Doc onward.
            if self._looks_like_resume_doc_text(app_status_value) and ai_fit_value in yes_no_values:
                corrected = list(row_padded)
                normalized_status = self._normalize_application_status(
                    app_status_value=app_status_value,
                    tail_status_value=tail_status_value,
                )

                for idx in range(target_len - 1, resume_doc_index, -1):
                    corrected[idx] = row_padded[idx - 1]

                corrected[resume_doc_index] = app_status_value
                corrected[app_status_index] = normalized_status
                repaired_rows.append(corrected)
                changed = True
                continue

            # Case 2: Legacy rows from pre-Application-Status schema.
            looks_like_old_layout = (
                ai_fit_value not in yes_no_values
                and (
                    job_summary_value in yes_no_values
                    or self._looks_like_resume_doc_text(ai_fit_value)
                    or (
                        self._looks_like_fit_score(resume_doc_value)
                        and self._looks_like_recommendation(fit_score_value)
                    )
                )
            )
            if looks_like_old_layout:
                corrected = [""] * target_len
                normalized_status = self._normalize_application_status(
                    app_status_value=app_status_value,
                    tail_status_value=tail_status_value,
                )
                col3_status = self._normalize_status_value(row_padded[2])

                corrected[0] = row_padded[0]
                corrected[1] = row_padded[1]
                corrected[app_status_index] = normalized_status
                corrected[3] = row_padded[2] if not col3_status else ""

                for new_idx in range(4, target_len):
                    corrected[new_idx] = row_padded[new_idx - 1]

                repaired_rows.append(corrected)
                changed = True
                continue

            repaired_rows.append(row_padded)

        if not changed:
            return

        end_col = self._column_index_to_letter(target_len)
        worksheet.update(
            range_name=f"A2:{end_col}{len(used_values)}",
            values=repaired_rows,
            value_input_option="USER_ENTERED",
        )

    @staticmethod
    def _looks_like_resume_doc_text(value: str) -> bool:
        if not value:
            return False
        lowered = value.lower()
        return value.startswith("Resume - ") or "/applications/" in lowered or lowered.endswith(".txt")

    @staticmethod
    def _looks_like_fit_score(value: str) -> bool:
        if not value:
            return False
        try:
            parsed = float(value)
        except ValueError:
            return False
        return 0 <= parsed <= 100

    @staticmethod
    def _looks_like_recommendation(value: str) -> bool:
        return value.strip().lower() in {"strong_match", "good_match", "stretch", "skip"}

    @classmethod
    def _normalize_status_value(cls, value: str) -> str:
        mapping = {
            "applied": "Applied",
            "interview": "Interview",
            "rejected": "Rejected",
            "accepted": "Accepted",
            "accept": "Accepted",
        }
        return mapping.get(value.strip().lower(), "")

    @classmethod
    def _normalize_application_status(cls, app_status_value: str, tail_status_value: str) -> str:
        from_status_col = cls._normalize_status_value(app_status_value)
        if from_status_col:
            return from_status_col
        from_tail_col = cls._normalize_status_value(tail_status_value)
        if from_tail_col:
            return from_tail_col
        return ""


class RemoteJobsWorksheetExporter(JobAutomationWorksheetExporter):
    @property
    def worksheet_name(self) -> str:
        return self.settings.google_sheets_remote_jobs_worksheet


class PhdRoleWorksheetExporter(BaseWorksheetExporter):
    @property
    def worksheet_name(self) -> str:
        return self.settings.google_sheets_phd_report_worksheet

    @property
    def headers(self) -> list[str]:
        return DEFAULT_PHD_REPORT_HEADERS
